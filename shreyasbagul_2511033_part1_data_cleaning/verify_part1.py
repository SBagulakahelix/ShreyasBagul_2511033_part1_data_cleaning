import os
import pandas as pd
import openpyxl
import re

print("--- Starting Automated Verification ---")

project_dir = r"C:\Users\bagul\OneDrive\Desktop\assignment\shreyasbagul_2511033_part1_data_cleaning"

# 1. Check Directory and Files Existence
expected_files = [
    os.path.join(project_dir, "data", "raw_orders.xlsx"),
    os.path.join(project_dir, "data", "cleaned_orders.xlsx"),
    os.path.join(project_dir, "outputs", "data_quality_report.xlsx"),
    os.path.join(project_dir, "outputs", "pivot_summary.xlsx"),
    os.path.join(project_dir, "outputs", "cleaning_log.md"),
    os.path.join(project_dir, "screenshots", "data_quality_report_preview.png"),
    os.path.join(project_dir, "screenshots", "cleaned_data_preview.png"),
    os.path.join(project_dir, "screenshots", "pivot_summary_preview1.png"),
    os.path.join(project_dir, "screenshots", "pivot_summary_preview2.png"),
    os.path.join(project_dir, "README.md")
]

print("\nChecking file structure:")
all_files_exist = True
for f in expected_files:
    if os.path.exists(f):
        size = os.path.getsize(f)
        print(f"  [PASS] {os.path.basename(f)} exists ({size} bytes)")
    else:
        print(f"  [FAIL] {f} does NOT exist!")
        all_files_exist = False

if not all_files_exist:
    print("Verification failed due to missing files!")
    exit(1)

# 2. Check Cleaned Orders Excel content
print("\nChecking cleaned_orders.xlsx:")
try:
    df_clean = pd.read_excel(os.path.join(project_dir, "data", "cleaned_orders.xlsx"), sheet_name="cleaned_orders")
    print(f"  [PASS] Load successful. Row count: {len(df_clean)} (Expected: 912)")
    if len(df_clean) != 912:
        print(f"  [FAIL] Row count is {len(df_clean)}, expected 912.")
except Exception as e:
    print(f"  [FAIL] Load failed: {e}")
    exit(1)

# Check Columns
required_cols = [
    'cleaned_discount', 'calculated_sales', 'calculated_profit', 'profit_margin',
    'shipping_delay_days', 'order_month', 'order_year', 'data_quality_flag'
]
cols_pass = True
for col in required_cols:
    if col in df_clean.columns:
        print(f"  [PASS] Column '{col}' exists")
    else:
        print(f"  [FAIL] Column '{col}' is missing!")
        cols_pass = False

# Check Text Casing & Trimming
text_columns = ['segment', 'region', 'category', 'sub_category', 'ship_mode', 'payment_status', 'order_status']
casing_pass = True
for col in text_columns:
    unique_vals = df_clean[col].dropna().unique()
    for val in unique_vals:
        if val != str(val).strip() or '  ' in str(val) or val != str(val).title():
            print(f"  [FAIL] Text format issue in '{col}': {repr(val)}")
            casing_pass = False
            break
if casing_pass:
    print("  [PASS] Text casing and spacing standardized successfully")

# Check Dates
date_format_pass = True
date_regex = re.compile(r'^\d{4}-\d{2}-\d{2}$')
for col in ['order_date', 'ship_date']:
    for idx, val in enumerate(df_clean[col]):
        if not date_regex.match(str(val)):
            print(f"  [FAIL] Date format issue in '{col}' row {idx}: {repr(val)}")
            date_format_pass = False
            break
if date_format_pass:
    print("  [PASS] All dates standardized to YYYY-MM-DD")

# 3. Check Data Quality Report Sheets
print("\nChecking data_quality_report.xlsx sheets:")
try:
    wb_dq = openpyxl.load_workbook(os.path.join(project_dir, "outputs", "data_quality_report.xlsx"), read_only=True)
    expected_dq_sheets = ["Summary_Dashboard", "Missing_Values", "Duplicate_Records", "Discount_Anomalies", "Date_Anomalies", "Calculation_Mismatches"]
    for s in expected_dq_sheets:
        if s in wb_dq.sheetnames:
            print(f"  [PASS] Sheet '{s}' exists")
        else:
            print(f"  [FAIL] Sheet '{s}' is missing!")
except Exception as e:
    print(f"  [FAIL] Load failed: {e}")

# 4. Check Pivot Summary Sheets
print("\nChecking pivot_summary.xlsx sheets:")
try:
    wb_pv = openpyxl.load_workbook(os.path.join(project_dir, "outputs", "pivot_summary.xlsx"), read_only=True)
    expected_pv_sheets = ["Sales_by_Region", "Sales_by_Category", "Orders_by_Ship_Mode", "Margin_by_Segment", "Issues_by_Region", "Monthly_Sales_Trend"]
    for s in expected_pv_sheets:
        if s in wb_pv.sheetnames:
            print(f"  [PASS] Sheet '{s}' exists")
        else:
            print(f"  [FAIL] Sheet '{s}' is missing!")
except Exception as e:
    print(f"  [FAIL] Load failed: {e}")

print("\n--- Verification Complete ---")
