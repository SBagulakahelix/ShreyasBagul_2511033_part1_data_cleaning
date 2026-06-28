import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# Define file paths
project_dir = r"C:\Users\bagul\OneDrive\Desktop\assignment\shreyasbagul_2511033_part1_data_cleaning"
clean_path = os.path.join(project_dir, "data", "cleaned_orders.xlsx")
pivot_path = os.path.join(project_dir, "outputs", "pivot_summary.xlsx")

# Load cleaned dataset
df = pd.read_excel(clean_path, sheet_name="cleaned_orders")

print("Generating Pivot Summary Report...")

# Create Excel Workbook
wb = openpyxl.Workbook()
# remove default sheet
default_sheet = wb.active
wb.remove(default_sheet)

# Style Definitions
font_family = "Segoe UI"
title_font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
subtitle_font = Font(name=font_family, size=10, italic=True, color="FFFFFF")
header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
data_font = Font(name=font_family, size=10)
bold_font = Font(name=font_family, size=10, bold=True)
total_font = Font(name=font_family, size=10, bold=True, color="1F4E79")

# Colors (Hex)
title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Deep steel blue
header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid") # Medium steel blue
total_fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid") # Light grey for totals
zebra_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid") # Zebra stripe

# Borders
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
total_border = Border(
    top=Side(style='thin', color='1F4E79'),
    bottom=Side(style='double', color='1F4E79'),
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9')
)

# Helper function to auto-fit columns and enable gridlines
def format_sheet_layout(ws, freeze_row=4):
    ws.views.sheetView[0].showGridLines = True
    if freeze_row > 0:
        ws.freeze_panes = f"A{freeze_row}"
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if cell.number_format and '%' in cell.number_format:
                val += '%'
            if '\n' in val:
                lines = val.split('\n')
                max_len = max(max_len, max(len(l) for l in lines))
            else:
                max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

# Helper function to write title block
def write_title_block(ws, title, subtitle):
    ws.merge_cells("A1:E1")
    ws.merge_cells("A2:E2")
    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"] = subtitle
    ws["A2"].font = subtitle_font
    ws["A2"].fill = title_fill
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 18

# Helper to write headers
def write_headers(ws, headers, start_row=4):
    ws.row_dimensions[start_row].height = 22
    for c_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=c_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Helper to style data row
def style_row(ws, row_idx, num_cols, align_left_cols=[1], number_formats={}):
    ws.row_dimensions[row_idx].height = 20
    is_zebra = (row_idx % 2 == 0)
    for c_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=c_idx)
        cell.font = data_font
        cell.border = thin_border
        if is_zebra:
            cell.fill = zebra_fill
            
        # Alignment
        if c_idx in align_left_cols:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            
        # Formatting
        if c_idx in number_formats:
            cell.number_format = number_formats[c_idx]

# Helper to write total row
def write_total_row(ws, row_idx, num_cols, label_col=1, formula_dict={}):
    ws.row_dimensions[row_idx].height = 22
    for c_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=c_idx)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = total_border
        
        if c_idx == label_col:
            cell.value = "Grand Total"
            cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx in formula_dict:
            cell.value = formula_dict[c_idx]
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            cell.value = ""

# Filter completed orders (Completed & Paid)
df_completed = df[(df['order_status'] == 'Completed') & (df['payment_status'] == 'Paid')].copy()

# ==========================================
# PIVOT 1: Sales & Profit by Region
# ==========================================
print("  - Sales_by_Region")
ws_reg = wb.create_sheet("Sales_by_Region")
write_title_block(ws_reg, "SALES & PROFIT BY REGION", "Completed Sales Only")
headers_reg = ["Region", "Sales", "Profit", "Profit Margin"]
write_headers(ws_reg, headers_reg)

# Aggregate and sort by sales descending
df_reg = df_completed.groupby('region')[['calculated_sales', 'calculated_profit']].sum().reset_index()
df_reg = df_reg.sort_values(by='calculated_sales', ascending=False)

for idx, row in enumerate(df_reg.values, 5):
    ws_reg.cell(row=idx, column=1, value=row[0])
    ws_reg.cell(row=idx, column=2, value=row[1])
    ws_reg.cell(row=idx, column=3, value=row[2])
    # Margin formula: Profit / Sales
    ws_reg.cell(row=idx, column=4, value=f"=C{idx}/B{idx}")
    
    style_row(ws_reg, idx, 4, align_left_cols=[1], number_formats={2: "$#,##0.00", 3: "$#,##0.00", 4: "0.0%"})

total_row_idx = 5 + len(df_reg)
write_total_row(ws_reg, total_row_idx, 4, label_col=1, formula_dict={
    2: f"=SUM(B5:B{total_row_idx-1})",
    3: f"=SUM(C5:C{total_row_idx-1})",
    4: f"=C{total_row_idx}/B{total_row_idx}"
})
ws_reg.cell(row=total_row_idx, column=2).number_format = "$#,##0.00"
ws_reg.cell(row=total_row_idx, column=3).number_format = "$#,##0.00"
ws_reg.cell(row=total_row_idx, column=4).number_format = "0.0%"

format_sheet_layout(ws_reg, freeze_row=5)

# ==========================================
# PIVOT 2: Sales & Profit by Category
# ==========================================
print("  - Sales_by_Category")
ws_cat = wb.create_sheet("Sales_by_Category")
write_title_block(ws_cat, "SALES & PROFIT BY CATEGORY & SUB-CATEGORY", "Completed Sales Only")
headers_cat = ["Category", "Sub-Category", "Sales", "Profit", "Profit Margin"]
write_headers(ws_cat, headers_cat)

# Aggregate and sort by category, and sub-category sales descending
df_cat = df_completed.groupby(['category', 'sub_category'])[['calculated_sales', 'calculated_profit']].sum().reset_index()
df_cat = df_cat.sort_values(by=['category', 'calculated_sales'], ascending=[True, False])

for idx, row in enumerate(df_cat.values, 5):
    ws_cat.cell(row=idx, column=1, value=row[0])
    ws_cat.cell(row=idx, column=2, value=row[1])
    ws_cat.cell(row=idx, column=3, value=row[2])
    ws_cat.cell(row=idx, column=4, value=row[3])
    ws_cat.cell(row=idx, column=5, value=f"=D{idx}/C{idx}")
    
    style_row(ws_cat, idx, 5, align_left_cols=[1, 2], number_formats={3: "$#,##0.00", 4: "$#,##0.00", 5: "0.0%"})

total_row_idx = 5 + len(df_cat)
write_total_row(ws_cat, total_row_idx, 5, label_col=1, formula_dict={
    3: f"=SUM(C5:C{total_row_idx-1})",
    4: f"=SUM(D5:D{total_row_idx-1})",
    5: f"=D{total_row_idx}/C{total_row_idx}"
})
ws_cat.cell(row=total_row_idx, column=3).number_format = "$#,##0.00"
ws_cat.cell(row=total_row_idx, column=4).number_format = "$#,##0.00"
ws_cat.cell(row=total_row_idx, column=5).number_format = "0.0%"

format_sheet_layout(ws_cat, freeze_row=5)

# ==========================================
# PIVOT 3: Order Count by Ship Mode
# ==========================================
print("  - Orders_by_Ship_Mode")
ws_ship = wb.create_sheet("Orders_by_Ship_Mode")
write_title_block(ws_ship, "ORDER COUNT BY SHIP MODE", "Completed Sales Only")
headers_ship = ["Ship Mode", "Order Count", "Percentage of Total"]
write_headers(ws_ship, headers_ship)

# Aggregate and sort by count descending
df_ship = df_completed.groupby('ship_mode').size().reset_index(name='count')
df_ship = df_ship.sort_values(by='count', ascending=False)

for idx, row in enumerate(df_ship.values, 5):
    ws_ship.cell(row=idx, column=1, value=row[0])
    ws_ship.cell(row=idx, column=2, value=row[1])
    # Percentage formula: Count / Grand Total Count
    ws_ship.cell(row=idx, column=3, value=f"=B{idx}/B{5 + len(df_ship)}")
    
    style_row(ws_ship, idx, 3, align_left_cols=[1], number_formats={2: "#,##0", 3: "0.0%"})

total_row_idx = 5 + len(df_ship)
write_total_row(ws_ship, total_row_idx, 3, label_col=1, formula_dict={
    2: f"=SUM(B5:B{total_row_idx-1})",
    3: f"=SUM(C5:C{total_row_idx-1})"
})
ws_ship.cell(row=total_row_idx, column=2).number_format = "#,##0"
ws_ship.cell(row=total_row_idx, column=3).number_format = "0.0%"

format_sheet_layout(ws_ship, freeze_row=5)

# ==========================================
# PIVOT 4: Profit Margin by Segment
# ==========================================
print("  - Margin_by_Segment")
ws_seg = wb.create_sheet("Margin_by_Segment")
write_title_block(ws_seg, "PROFIT MARGIN BY CUSTOMER SEGMENT", "Completed Sales Only")
headers_seg = ["Segment", "Sales", "Profit", "Profit Margin"]
write_headers(ws_seg, headers_seg)

# Aggregate
df_seg = df_completed.groupby('segment')[['calculated_sales', 'calculated_profit']].sum().reset_index()
df_seg = df_seg.sort_values(by='calculated_sales', ascending=False)

for idx, row in enumerate(df_seg.values, 5):
    ws_seg.cell(row=idx, column=1, value=row[0])
    ws_seg.cell(row=idx, column=2, value=row[1])
    ws_seg.cell(row=idx, column=3, value=row[2])
    ws_seg.cell(row=idx, column=4, value=f"=C{idx}/B{idx}")
    
    style_row(ws_seg, idx, 4, align_left_cols=[1], number_formats={2: "$#,##0.00", 3: "$#,##0.00", 4: "0.0%"})

total_row_idx = 5 + len(df_seg)
write_total_row(ws_seg, total_row_idx, 4, label_col=1, formula_dict={
    2: f"=SUM(B5:B{total_row_idx-1})",
    3: f"=SUM(C5:C{total_row_idx-1})",
    4: f"=C{total_row_idx}/B{total_row_idx}"
})
ws_seg.cell(row=total_row_idx, column=2).number_format = "$#,##0.00"
ws_seg.cell(row=total_row_idx, column=3).number_format = "$#,##0.00"
ws_seg.cell(row=total_row_idx, column=4).number_format = "0.0%"

format_sheet_layout(ws_seg, freeze_row=5)

# ==========================================
# PIVOT 5: Refunded, Cancelled, Failed Orders by Region
# ==========================================
print("  - Issues_by_Region")
ws_iss = wb.create_sheet("Issues_by_Region")
write_title_block(ws_iss, "REFUNDED / CANCELLED / FAILED ORDERS BY REGION", "All Non-Completed or Flagged Orders")
headers_iss = ["Region", "Cancelled Count", "Returned Count", "Failed Payments", "Refunded Payments", "Total Issues"]
write_headers(ws_iss, headers_iss)

# Filter rows that represent issues (non-completed or failed/refunded payments)
# Cancelled: order_status == 'Cancelled'
# Returned: order_status == 'Returned'
# Failed: payment_status == 'Failed'
# Refunded: payment_status == 'Refunded'
df_issues = df[(df['order_status'] != 'Completed') | (df['payment_status'] != 'Paid')].copy()

# Group by region and get counts
regions_unique = sorted(df['region'].unique())
issue_counts = []

for reg in regions_unique:
    df_reg_iss = df_issues[df_issues['region'] == reg]
    cancelled_cnt = (df_reg_iss['order_status'] == 'Cancelled').sum()
    returned_cnt = (df_reg_iss['order_status'] == 'Returned').sum()
    failed_cnt = (df_reg_iss['payment_status'] == 'Failed').sum()
    refunded_cnt = (df_reg_iss['payment_status'] == 'Refunded').sum()
    
    issue_counts.append((reg, cancelled_cnt, returned_cnt, failed_cnt, refunded_cnt))

# Sort by region with most issues
issue_counts = sorted(issue_counts, key=lambda x: sum(x[1:]), reverse=True)

for idx, row in enumerate(issue_counts, 5):
    ws_iss.cell(row=idx, column=1, value=row[0])
    ws_iss.cell(row=idx, column=2, value=row[1])
    ws_iss.cell(row=idx, column=3, value=row[2])
    ws_iss.cell(row=idx, column=4, value=row[3])
    ws_iss.cell(row=idx, column=5, value=row[4])
    ws_iss.cell(row=idx, column=6, value=f"=SUM(B{idx}:E{idx})")
    
    style_row(ws_iss, idx, 6, align_left_cols=[1], number_formats={2: "#,##0", 3: "#,##0", 4: "#,##0", 5: "#,##0", 6: "#,##0"})

total_row_idx = 5 + len(issue_counts)
write_total_row(ws_iss, total_row_idx, 6, label_col=1, formula_dict={
    2: f"=SUM(B5:B{total_row_idx-1})",
    3: f"=SUM(C5:C{total_row_idx-1})",
    4: f"=SUM(D5:D{total_row_idx-1})",
    5: f"=SUM(E5:E{total_row_idx-1})",
    6: f"=SUM(F5:F{total_row_idx-1})"
})
for c in range(2, 7):
    ws_iss.cell(row=total_row_idx, column=c).number_format = "#,##0"

format_sheet_layout(ws_iss, freeze_row=5)

# ==========================================
# PIVOT 6: Monthly Sales Trend
# ==========================================
print("  - Monthly_Sales_Trend")
ws_trend = wb.create_sheet("Monthly_Sales_Trend")
write_title_block(ws_trend, "MONTHLY SALES AND PROFIT TREND", "Completed Sales Only")
headers_trend = ["Year", "Month", "Sales", "Profit", "Profit Margin"]
write_headers(ws_trend, headers_trend)

# Define month sorting order
month_order = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]

# Group by Year and Month
df_trend = df_completed.groupby(['order_year', 'order_month'])[['calculated_sales', 'calculated_profit']].sum().reset_index()
# Map month name to integer for sorting
df_trend['month_num'] = df_trend['order_month'].apply(lambda m: month_order.index(m) if m in month_order else 99)
df_trend = df_trend.sort_values(by=['order_year', 'month_num']).drop(columns=['month_num'])

for idx, row in enumerate(df_trend.values, 5):
    ws_trend.cell(row=idx, column=1, value=row[0])
    ws_trend.cell(row=idx, column=2, value=row[1])
    ws_trend.cell(row=idx, column=3, value=row[2])
    ws_trend.cell(row=idx, column=4, value=row[3])
    ws_trend.cell(row=idx, column=5, value=f"=D{idx}/C{idx}")
    
    style_row(ws_trend, idx, 5, align_left_cols=[1, 2], number_formats={3: "$#,##0.00", 4: "$#,##0.00", 5: "0.0%"})

total_row_idx = 5 + len(df_trend)
write_total_row(ws_trend, total_row_idx, 5, label_col=1, formula_dict={
    3: f"=SUM(C5:C{total_row_idx-1})",
    4: f"=SUM(D5:D{total_row_idx-1})",
    5: f"=D{total_row_idx}/C{total_row_idx}"
})
ws_trend.cell(row=total_row_idx, column=3).number_format = "$#,##0.00"
ws_trend.cell(row=total_row_idx, column=4).number_format = "$#,##0.00"
ws_trend.cell(row=total_row_idx, column=5).number_format = "0.0%"

format_sheet_layout(ws_trend, freeze_row=5)

# Save the workbook
print(f"Saving workbook to {pivot_path}...")
wb.save(pivot_path)
print("Pivot Summary Report generated successfully!")
