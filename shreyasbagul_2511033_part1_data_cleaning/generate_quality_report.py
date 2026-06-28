import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# Define file paths
project_dir = r"C:\Users\bagul\OneDrive\Desktop\assignment\shreyasbagul_2511033_part1_data_cleaning"
raw_path = os.path.join(project_dir, "data", "raw_orders.xlsx")
clean_path = os.path.join(project_dir, "data", "cleaned_orders.xlsx")
report_path = os.path.join(project_dir, "outputs", "data_quality_report.xlsx")

# Load datasets
df_raw = pd.read_excel(raw_path, sheet_name="raw_orders")
df_clean = pd.read_excel(clean_path, sheet_name="cleaned_orders")

# Calculate duplicate order ID series
dup_order_id_series = df_clean[df_clean.duplicated(subset=['order_id'], keep=False)]['order_id'].unique()

print("Generating Data Quality Report...")

# Create Excel Workbook
wb = openpyxl.Workbook()
# remove default sheet
default_sheet = wb.active
wb.remove(default_sheet)

# Style Definitions
font_family = "Segoe UI"
title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
subtitle_font = Font(name=font_family, size=11, italic=True, color="FFFFFF")
section_font = Font(name=font_family, size=12, bold=True, color="1F4E79")
header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
data_font = Font(name=font_family, size=10)
bold_font = Font(name=font_family, size=10, bold=True)
italic_font = Font(name=font_family, size=10, italic=True)

# Colors (Hex)
title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Deep steel blue
header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid") # Medium steel blue
section_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # Very light blue
zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Light gray
clean_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft green
warning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Soft yellow
invalid_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft orange/red

# Borders
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
header_border = Border(
    left=Side(style='thin', color='FFFFFF'),
    right=Side(style='thin', color='FFFFFF'),
    top=Side(style='thin', color='2F5597'),
    bottom=Side(style='medium', color='1F4E79')
)

# Helper function to auto-fit columns and enable gridlines
def format_sheet_layout(ws, freeze_row=4):
    ws.views.sheetView[0].showGridLines = True
    # Freeze panes
    if freeze_row > 0:
        ws.freeze_panes = f"A{freeze_row}"
    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if cell.number_format and '%' in cell.number_format:
                val += '%'
            if '\n' in val:
                lines = val.split('\n')
                max_len = max(max_len, max(len(l) for l in lines))
            else:
                max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

# Helper function to write stylized title block
def write_title_block(ws, title, subtitle):
    ws.merge_cells("A1:K1")
    ws.merge_cells("A2:K2")
    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"] = subtitle
    ws["A2"].font = subtitle_font
    ws["A2"].fill = title_fill
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20

# Helper function to write table headers
def write_table_headers(ws, headers, start_row):
    ws.row_dimensions[start_row].height = 25
    for c_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=c_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border

# ==========================================
# SHEET 1: Summary_Dashboard
# ==========================================
print("  - Summary_Dashboard")
ws_dash = wb.create_sheet("Summary_Dashboard")
write_title_block(ws_dash, "DATA QUALITY & CLEANING DASHBOARD", "Project: Shreyas Bagul (ID: 2511033) - Part 1")

# KPI blocks starting from row 4
kpis = [
    ("Total Raw Records", len(df_raw)),
    ("Exact Duplicates Removed", len(df_raw) - len(df_clean)),
    ("Clean Records (No Issues)", (df_clean['data_quality_flag'] == 'Clean').sum()),
    ("Warning Records (Imputed/Review)", (df_clean['data_quality_flag'] == 'Warning').sum()),
    ("Invalid Records (Flagged Error)", (df_clean['data_quality_flag'] == 'Invalid').sum())
]

for idx, (kpi_name, kpi_val) in enumerate(kpis):
    col_start = idx * 2 + 1
    col_end = col_start + 1
    col_letter_start = get_column_letter(col_start)
    col_letter_end = get_column_letter(col_end)
    
    ws_dash.merge_cells(f"{col_letter_start}4:{col_letter_end}4")
    ws_dash.merge_cells(f"{col_letter_start}5:{col_letter_end}5")
    
    cell_lbl = ws_dash[f"{col_letter_start}4"]
    cell_lbl.value = kpi_name
    cell_lbl.font = Font(name=font_family, size=9, bold=True, color="595959")
    cell_lbl.alignment = Alignment(horizontal="center", vertical="center")
    cell_lbl.fill = section_fill
    cell_lbl.border = thin_border
    
    cell_val = ws_dash[f"{col_letter_start}5"]
    cell_val.value = kpi_val
    cell_val.font = Font(name=font_family, size=14, bold=True, color="1F4E79")
    cell_val.alignment = Alignment(horizontal="center", vertical="center")
    cell_val.fill = section_fill
    cell_val.border = thin_border

ws_dash.row_dimensions[4].height = 18
ws_dash.row_dimensions[5].height = 25

# Summary Table 1: Record Distribution
ws_dash.cell(row=7, column=1, value="Data Quality Distribution").font = section_font
headers1 = ["Data Quality Flag", "Record Count", "Percentage", "Description / Action Taken"]
write_table_headers(ws_dash, headers1, 8)

dist_data = [
    ("Clean", (df_clean['data_quality_flag']=='Clean').sum(), "83.33%", "No anomalies detected; ready for final completed sales summary."),
    ("Warning", (df_clean['data_quality_flag']=='Warning').sum(), "11.07%", "Contains minor issues (missing region/ship_mode filled, sales mismatches, or conflicting duplicates). Included in analysis but marked for business review."),
    ("Invalid", (df_clean['data_quality_flag']=='Invalid').sum(), "5.59%", "Contains critical issues (negative/high discount, ship date before order date). Excluded from completed sales summaries.")
]

for r_idx, row_val in enumerate(dist_data, 9):
    ws_dash.row_dimensions[r_idx].height = 22
    for c_idx, val in enumerate(row_val, 1):
        cell = ws_dash.cell(row=r_idx, column=c_idx, value=val)
        cell.font = data_font
        cell.border = thin_border
        if c_idx == 1:
            cell.alignment = Alignment(horizontal="center")
            if val == "Clean": cell.fill = clean_fill
            elif val == "Warning": cell.fill = warning_fill
            elif val == "Invalid": cell.fill = invalid_fill
        elif c_idx == 2:
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
        elif c_idx == 3:
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.alignment = Alignment(horizontal="left")

# Summary Table 2: Detailed Anomaly Breakdown
ws_dash.cell(row=13, column=1, value="Detailed Anomaly Breakdown").font = section_font
headers2 = ["Anomaly / Rule Area", "Affected Rows", "Data Quality Flag", "Cleaning Action / Logic Applied"]
write_table_headers(ws_dash, headers2, 14)

anomaly_breakdown = [
    ("Exact Duplicate Rows", len(df_raw) - len(df_clean), "Removed", "Removed entirely from the cleaned dataset (20 exact duplicates deleted)."),
    ("Conflicting Duplicate Order IDs", 24, "Warning", "Kept both conflicting copies in cleaned dataset; flagged data_quality_flag = 'Warning' for review."),
    ("Missing Region", 25, "Warning", "Imputed with value 'Unknown'. Flagged in quality report."),
    ("Missing Ship Mode", 21, "Warning", "Imputed with value 'Unknown'. Flagged in quality report."),
    ("Missing Discount (Math Matches)", 4, "Clean", "Treated missing discount as 0.0 because quantity * unit_price equals sales. No flag needed."),
    ("Missing Discount (Math Mismatch)", 14, "Warning", "Treated as 0.0 but flagged as Warning due to sales mismatch (indicates discount was applied but unrecorded)."),
    ("Negative Discount", 15, "Invalid", "Flagged discount < 0 (e.g. -0.15) as Invalid row."),
    ("Discount Above Allowed Range", 15, "Invalid", "Flagged discount > 25% (unusually high, e.g. 55%, 65%, 70%, 85%) as Invalid row."),
    ("Ship Date Earlier Than Order Date", 21, "Invalid", "Flagged invalid sequence (shipping delay negative, -1 to -4 days) as Invalid row."),
    ("Sales/Profit Recalculation Mismatch", 64, "Warning", "Flagged rows where raw sales/profit differs from recalculated value by > $0.05.")
]

for r_idx, row_val in enumerate(anomaly_breakdown, 15):
    ws_dash.row_dimensions[r_idx].height = 22
    for c_idx, val in enumerate(row_val, 1):
        cell = ws_dash.cell(row=r_idx, column=c_idx, value=val)
        cell.font = data_font
        cell.border = thin_border
        if c_idx == 1:
            cell.alignment = Alignment(horizontal="left")
        elif c_idx == 2:
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
        elif c_idx == 3:
            cell.alignment = Alignment(horizontal="center")
            if val == "Clean" or val == "Removed": cell.fill = clean_fill
            elif val == "Warning": cell.fill = warning_fill
            elif val in ["Invalid", "Removed (Flagged)"]: cell.fill = invalid_fill
        else:
            cell.alignment = Alignment(horizontal="left", wrap_text=True)

format_sheet_layout(ws_dash, freeze_row=0)

# ==========================================
# Helper to write raw data sheets
# ==========================================
def write_detail_sheet(sheet_name, df_slice, title, headers, extra_formatting_func=None):
    print(f"  - {sheet_name}")
    ws = wb.create_sheet(sheet_name)
    write_title_block(ws, title, "Data Quality Report Details - Shreyas Bagul (2511033)")
    write_table_headers(ws, headers, 4)
    
    # write rows
    for r_idx, row in enumerate(df_slice.values, 5):
        ws.row_dimensions[r_idx].height = 20
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            # formatting numbers
            if isinstance(val, float):
                if abs(val) <= 1.0 and (headers[c_idx-1] in ['discount', 'cleaned_discount', 'profit_margin', 'Discount', 'Cleaned Discount', 'Profit Margin']):
                    cell.number_format = "0.0%"
                else:
                    cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif isinstance(val, (int, np.integer)):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            elif pd.isnull(val):
                cell.value = "-"
                cell.alignment = Alignment(horizontal="center")
            elif str(val) in ["Clean", "Warning", "Invalid"]:
                cell.alignment = Alignment(horizontal="center")
                if val == "Clean": cell.fill = clean_fill
                elif val == "Warning": cell.fill = warning_fill
                elif val == "Invalid": cell.fill = invalid_fill
            else:
                cell.alignment = Alignment(horizontal="left")
                
    if extra_formatting_func:
        extra_formatting_func(ws, len(df_slice))
        
    format_sheet_layout(ws, freeze_row=5)

# ==========================================
# SHEET 2: Missing_Values
# ==========================================
# Columns: order_id, customer_name, region, ship_mode, discount, sales, cost, profit, order_status, data_quality_flag
# We want to show rows that had missing region, ship_mode, or discount in df_raw.
missing_raw_ids = df_raw[df_raw['region'].isnull() | df_raw['ship_mode'].isnull() | df_raw['discount'].isnull()]['order_id'].unique()
df_clean_missing = df_clean[df_clean['order_id'].isin(missing_raw_ids)][[
    'order_id', 'customer_name', 'region', 'ship_mode', 'discount', 'cleaned_discount', 'sales', 'cost', 'profit', 'order_status', 'data_quality_flag'
]]

headers_missing = ["Order ID", "Customer Name", "Region (Imputed)", "Ship Mode (Imputed)", "Raw Discount", "Cleaned Discount", "Sales", "Cost", "Profit", "Order Status", "DQ Flag"]

def highlight_missing_imputations(ws, row_count):
    # Highlight imputed values in cells
    for r in range(5, 5 + row_count):
        # check region (col 3)
        region_cell = ws.cell(row=r, column=3)
        if region_cell.value == "Unknown":
            region_cell.fill = warning_fill
            region_cell.font = bold_font
            
        # check ship_mode (col 4)
        ship_mode_cell = ws.cell(row=r, column=4)
        if ship_mode_cell.value == "Unknown":
            ship_mode_cell.fill = warning_fill
            ship_mode_cell.font = bold_font
            
        # check discount (col 5 is raw discount)
        raw_disc_cell = ws.cell(row=r, column=5)
        clean_disc_cell = ws.cell(row=r, column=6)
        if raw_disc_cell.value == "-":
            # imputed missing discount
            clean_disc_cell.fill = warning_fill
            clean_disc_cell.font = bold_font

write_detail_sheet("Missing_Values", df_clean_missing, "IMPUTED MISSING VALUES SUMMARY", headers_missing, highlight_missing_imputations)

# ==========================================
# SHEET 3: Duplicate_Records
# ==========================================
# Conflicting duplicates (both copies of the 12 order IDs)
df_clean_dups = df_clean[df_clean['order_id'].isin(dup_order_id_series)][[
    'order_id', 'order_date', 'ship_date', 'customer_name', 'sales', 'calculated_sales', 'cost', 'profit', 'order_status', 'payment_status', 'data_quality_flag'
]].sort_values('order_id')

headers_dups = ["Order ID", "Order Date", "Ship Date", "Customer Name", "Raw Sales", "Calculated Sales", "Cost", "Profit", "Order Status", "Payment Status", "DQ Flag"]

def format_duplicates_sheet(ws, row_count):
    # we want to highlight conflicting duplicate rows
    for r in range(5, 5 + row_count):
        ws.cell(row=r, column=11).fill = warning_fill
        ws.cell(row=r, column=11).font = bold_font

write_detail_sheet("Duplicate_Records", df_clean_dups, "CONFLICTING DUPLICATE ORDER IDS SUMMARY", headers_dups, format_duplicates_sheet)

# ==========================================
# SHEET 4: Discount_Anomalies
# ==========================================
# Negative discounts or discounts > 0.25
df_clean_disc_anom = df_clean[(df_clean['cleaned_discount'] < 0) | (df_clean['cleaned_discount'] > 0.25)][[
    'order_id', 'quantity', 'unit_price', 'discount', 'cleaned_discount', 'sales', 'calculated_sales', 'cost', 'profit', 'data_quality_flag'
]]

headers_disc_anom = ["Order ID", "Quantity", "Unit Price", "Raw Discount", "Cleaned Discount", "Raw Sales", "Calculated Sales", "Cost", "Profit", "DQ Flag"]

def format_discount_anom_sheet(ws, row_count):
    for r in range(5, 5 + row_count):
        cell_disc = ws.cell(row=r, column=5)
        cell_flag = ws.cell(row=r, column=10)
        # invalid color
        cell_disc.fill = invalid_fill
        cell_disc.font = bold_font
        cell_flag.fill = invalid_fill
        cell_flag.font = bold_font

write_detail_sheet("Discount_Anomalies", df_clean_disc_anom, "DISCOUNT ANOMALIES SUMMARY (NEGATIVE & ABOVE 25%)", headers_disc_anom, format_discount_anom_sheet)

# ==========================================
# SHEET 5: Date_Anomalies
# ==========================================
# Ship date < order date
df_clean_date_anom = df_clean[df_clean['shipping_delay_days'] < 0][[
    'order_id', 'order_date', 'ship_date', 'shipping_delay_days', 'ship_mode', 'order_status', 'payment_status', 'data_quality_flag'
]]

headers_date_anom = ["Order ID", "Order Date", "Ship Date", "Shipping Delay (Days)", "Ship Mode", "Order Status", "Payment Status", "DQ Flag"]

def format_date_anom_sheet(ws, row_count):
    for r in range(5, 5 + row_count):
        cell_delay = ws.cell(row=r, column=4)
        cell_flag = ws.cell(row=r, column=8)
        cell_delay.fill = invalid_fill
        cell_delay.font = bold_font
        cell_flag.fill = invalid_fill
        cell_flag.font = bold_font

write_detail_sheet("Date_Anomalies", df_clean_date_anom, "DATE SEQUENCE ANOMALIES SUMMARY (SHIP DATE < ORDER DATE)", headers_date_anom, format_date_anom_sheet)

# ==========================================
# SHEET 6: Calculation_Mismatches
# ==========================================
# Sales or profit mismatches
df_clean['sales_diff'] = df_clean['sales'] - df_clean['calculated_sales']
df_clean['profit_diff'] = df_clean['profit'] - df_clean['calculated_profit']

df_clean_math_mismatch = df_clean[
    (df_clean['sales_diff'].abs() > 0.05) | (df_clean['profit_diff'].abs() > 0.05)
][[
    'order_id', 'quantity', 'unit_price', 'cleaned_discount', 'sales', 'calculated_sales', 'sales_diff', 'profit', 'calculated_profit', 'profit_diff', 'data_quality_flag'
]]

headers_math = [
    "Order ID", "Qty", "Unit Price", "Discount", "Raw Sales", "Calculated Sales", "Sales Diff", "Raw Profit", "Calculated Profit", "Profit Diff", "DQ Flag"
]

def format_math_sheet(ws, row_count):
    for r in range(5, 5 + row_count):
        # col 7 is Sales Diff, col 10 is Profit Diff
        for col_idx in [7, 10]:
            cell = ws.cell(row=r, column=col_idx)
            cell.fill = warning_fill
            cell.font = bold_font

write_detail_sheet("Calculation_Mismatches", df_clean_math_mismatch, "SALES & PROFIT RECALCULATION MISMATCH SUMMARY", headers_math, format_math_sheet)

# Save the workbook
print(f"Saving workbook to {report_path}...")
wb.save(report_path)
print("Data Quality Report generated successfully!")
